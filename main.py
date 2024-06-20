import openpyxl as xl

def xls(filename):
    # load workbook
    wb = xl.load_workbook(filename)
    # select the sheet number.
    sheet = wb["Sheet1"]
    total_rate = 0

    for row in range(2, sheet.max_row ):
        cell1 = sheet.cell(row, 3)
        cell2 = sheet.cell(row, 4)
        if cell1.value  and cell2.value is not None:
            try:
                rate = float(cell1.value) * float(cell2.value)
                total_rate += rate
            except ValueError:
                rate = None  # Handle the case where conversion to float fails
        else:
            rate = None  # Handle the case where one of the values is None
        corrected_cell = sheet.cell(row, 5 )
        corrected_cell.value = rate

    total_rate_cell = sheet.cell(12, 5)
    total_rate_cell.value = total_rate

    wb.save(filename)


xls("Book1.xlsx")
xls("Book2.xlsx")
xls("Book3.xlsx")

'''
# for the indentation I use
# this paragraph is about you addressing my effort
'''